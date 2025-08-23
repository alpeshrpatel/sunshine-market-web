# from fastapi import APIRouter, UploadFile, File, Form
# from fastapi.responses import FileResponse
# from app.api.controllers.invoice_controller import process_invoice

# router = APIRouter()

# @router.post("/analyze-and-update")
# async def analyze_and_update_invoice(
#     vendor: str = Form(...),
#     pdf_file: UploadFile = File(...),
#     excel_file: UploadFile = File(...)
# ):
#     return await process_invoice(pdf_file, excel_file, vendor)


# @router.get("/download-updated")
# async def download_updated():
#     path = "updated_invoice.xlsx"
#     return FileResponse(
#         path=path,
#         filename="updated_invoice.xlsx",
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={
#             "Cache-Control": "no-store",
#             "Content-Disposition": "attachment; filename=updated_invoice.xlsx",
#             "Access-Control-Allow-Origin": "*"
#         }
#     )



# from fastapi import APIRouter, UploadFile, File, Form
# from fastapi.responses import FileResponse, JSONResponse
# import os
# import uuid
# from app.api.controllers.invoice_controller import process_invoice

# router = APIRouter()

# # Store processed files temporarily
# processed_files = {}

# @router.post("/analyze-and-update")
# async def analyze_and_update_invoice(
#     vendor: str = Form(...),
#     pdf_file: UploadFile = File(...),
#     excel_file: UploadFile = File(...)
# ):
#     try:
#         # Process the invoice
#         result = await process_invoice(pdf_file, excel_file, vendor)
        
#         # Generate a unique file ID
#         file_id = str(uuid.uuid4())
        
#         # Store the file path temporarily (you might want to use Redis or database in production)
#         processed_files[file_id] = {
#             "file_path": "updated_invoice.xlsx",  # Your processed file path
#             "filename": "updated_invoice.xlsx"
#         }
        
#         # Return JSON response instead of FileResponse
#         return JSONResponse({
#             "status": "success",
#             "message": "Invoice processed successfully",
#             "excel_data": result.get("excel_data", {}),
#             "download_url": f"/api/invoice/download/{file_id}"
#         })
        
#     except Exception as e:
#         return JSONResponse(
#             status_code=500,
#             content={
#                 "status": "error", 
#                 "message": f"Processing failed: {str(e)}"
#             }
#         )

# @router.get("/download/{file_id}")
# async def download_file(file_id: str):
#     """Download endpoint that uses file ID to retrieve the processed file"""
#     if file_id not in processed_files:
#         return JSONResponse(
#             status_code=404,
#             content={"status": "error", "message": "File not found or expired"}
#         )
    
#     file_info = processed_files[file_id]
#     file_path = file_info["file_path"]
#     filename = file_info["filename"]
    
#     # Check if file exists
#     if not os.path.exists(file_path):
#         return JSONResponse(
#             status_code=404,
#             content={"status": "error", "message": "File not found on server"}
#         )
    
#     # Clean up the temporary reference
#     del processed_files[file_id]
    
#     return FileResponse(
#         path=file_path,
#         filename=filename,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={
#             "Cache-Control": "no-store",
#             "Content-Disposition": f"attachment; filename={filename}",
#             "Access-Control-Allow-Origin": "*",
#             "Access-Control-Allow-Headers": "*"
#         }
#     )

# # Alternative approach: Return file as base64 in JSON (for smaller files)
# @router.post("/analyze-and-update-with-file")
# async def analyze_and_update_with_file(
#     vendor: str = Form(...),
#     pdf_file: UploadFile = File(...),
#     excel_file: UploadFile = File(...)
# ):
#     try:
#         # Process the invoice
#         result = await process_invoice(pdf_file, excel_file, vendor)
        
#         # Read the processed file and encode as base64
#         import base64
#         with open("updated_invoice.xlsx", "rb") as f:
#             file_content = base64.b64encode(f.read()).decode()
        
#         return JSONResponse({
#             "status": "success",
#             "message": "Invoice processed successfully",
#             "excel_data": result.get("excel_data", {}),
#             "file_content": file_content,
#             "filename": "updated_invoice.xlsx"
#         })
        
#     except Exception as e:
#         return JSONResponse(
#             status_code=500,
#             content={
#                 "status": "error", 
#                 "message": f"Processing failed: {str(e)}"
#             }
#         )









# routes.py
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse
import os
import uuid
import tempfile
import shutil
from pathlib import Path
from app.api.controllers.invoice_controller import process_invoice

router = APIRouter()

# Store processed files temporarily with better structure
processed_files = {}

@router.post("/analyze-and-update")
async def analyze_and_update_invoice(
    vendor: str = Form(...),
    pdf_file: UploadFile = File(...),
    excel_file: UploadFile = File(...)
):
    """
    Process invoice and return JSON response with download link
    """
    temp_dir = None
    try:
        print(f"🔄 Starting invoice analysis for vendor: {vendor}")
        
        # Create temporary directory for this request
        temp_dir = tempfile.mkdtemp()
        
        # Process the invoice with temp directory
        result = await process_invoice(pdf_file, excel_file, vendor, temp_dir)
        
        # Check if processing was successful
        if result.get("status") != "success":
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "message": result.get("message", "Processing failed"),
                    "excel_data": []
                }
            )
        
        # Generate a unique file ID
        file_id = str(uuid.uuid4())
        
        # Store the file info with absolute path
        output_file = os.path.join(temp_dir, "updated_invoice.xlsx")
        
        if not os.path.exists(output_file):
            return JSONResponse(
                status_code=500,
                content={
                    "status": "error",
                    "message": "Processed file was not created",
                    "excel_data": []
                }
            )
        
        processed_files[file_id] = {
            "file_path": output_file,
            "filename": f"updated_invoice_{vendor}.xlsx",
            "temp_dir": temp_dir,
            "created_at": str(uuid.uuid4())  # For cleanup tracking
        }
        
        print(f"✅ File processed successfully. ID: {file_id}")
        
        # Return JSON response
        return JSONResponse(
            status_code=200,
            content={
                "status": "success",
                "message": f"Invoice processed successfully. Updated {result.get('processed_skus', 0)} SKUs.",
                "excel_data": result.get("excel_data", []),
                "download_url": f"/api/invoice/download/{file_id}",
                "file_id": file_id,
                "processed_skus": result.get("processed_skus", 0)
            }
        )
        
    except Exception as e:
        print(f"❌ Error in analyze_and_update_invoice: {str(e)}")
        
        # Clean up temp directory on error
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception as cleanup_error:
                print(f"⚠️ Cleanup error: {cleanup_error}")
        
        return JSONResponse(
            status_code=500,
            content={
                "status": "error",
                "message": f"Processing failed: {str(e)}",
                "excel_data": []
            }
        )

@router.get("/download/{file_id}")
async def download_file(file_id: str):
    """
    Download endpoint that uses file ID to retrieve the processed file
    """
    try:
        if file_id not in processed_files:
            raise HTTPException(
                status_code=404,
                detail="File not found or expired"
            )
        
        file_info = processed_files[file_id]
        file_path = file_info["file_path"]
        filename = file_info["filename"]
        temp_dir = file_info.get("temp_dir")
        
        # Check if file exists
        if not os.path.exists(file_path):
            # Clean up the reference
            if file_id in processed_files:
                del processed_files[file_id]
            
            raise HTTPException(
                status_code=404,
                detail="File not found on server"
            )
        
        print(f"📥 Downloading file: {filename} for ID: {file_id}")
        
        # Create a response with proper headers
        response = FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Add headers to prevent caching and handle CORS
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        response.headers["Content-Disposition"] = f"attachment; filename=\"{filename}\""
        
        # Clean up after successful response creation
        # Note: Don't delete immediately, let the response finish first
        def cleanup():
            try:
                if file_id in processed_files:
                    del processed_files[file_id]
                if temp_dir and os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
            except Exception as e:
                print(f"⚠️ Cleanup error: {e}")
        
        # Schedule cleanup (you might want to implement this differently)
        # For now, we'll clean up after a delay or implement a background task
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error in download_file: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Download failed: {str(e)}"
        )

@router.post("/analyze-and-update-inline")
async def analyze_and_update_inline(
    vendor: str = Form(...),
    pdf_file: UploadFile = File(...),
    excel_file: UploadFile = File(...)
):
    """
    Process invoice and return file content as base64 in JSON response
    Better for smaller files and avoids download dialog issues
    """
    temp_dir = None
    try:
        print(f"🔄 Starting inline invoice analysis for vendor: {vendor}")
        
        # Create temporary directory for this request
        temp_dir = tempfile.mkdtemp()
        
        # Process the invoice
        result = await process_invoice(pdf_file, excel_file, vendor, temp_dir)
        
        # Check if processing was successful
        if result.get("status") != "success":
            return JSONResponse(
                status_code=400,
                content={
                    "status": "error",
                    "message": result.get("message", "Processing failed"),
                    "excel_data": []
                }
            )
        
        # Read the processed file and encode as base64
        output_file = os.path.join(temp_dir, "updated_invoice.xlsx")
        
        if not os.path.exists(output_file):
            return JSONResponse(
                status_code=500,
                content={
                    "status": "error",
                    "message": "Processed file was not created",
                    "excel_data": []
                }
            )
        
        import base64
        with open(output_file, "rb") as f:
            file_content = base64.b64encode(f.read()).decode()
        
        # Clean up temp directory
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception as cleanup_error:
                print(f"⚠️ Cleanup error: {cleanup_error}")
        
        return JSONResponse(
            status_code=200,
            content={
                "status": "success",
                "message": f"Invoice processed successfully. Updated {result.get('processed_skus', 0)} SKUs.",
                "excel_data": result.get("excel_data", []),
                "file_content": file_content,
                "filename": f"updated_invoice_{vendor}.xlsx",
                "processed_skus": result.get("processed_skus", 0)
            }
        )
        
    except Exception as e:
        print(f"❌ Error in analyze_and_update_inline: {str(e)}")
        
        # Clean up temp directory on error
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception as cleanup_error:
                print(f"⚠️ Cleanup error: {cleanup_error}")
        
        return JSONResponse(
            status_code=500,
            content={
                "status": "error",
                "message": f"Processing failed: {str(e)}",
                "excel_data": []
            }
        )

@router.delete("/cleanup/{file_id}")
async def cleanup_file(file_id: str):
    """
    Cleanup endpoint to remove temporary files
    """
    try:
        if file_id in processed_files:
            file_info = processed_files[file_id]
            temp_dir = file_info.get("temp_dir")
            
            # Remove file reference
            del processed_files[file_id]
            
            # Clean up temp directory
            if temp_dir and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            
            return JSONResponse(
                content={"status": "success", "message": "File cleaned up"}
            )
        else:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "File not found"}
            )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"Cleanup failed: {str(e)}"}
        )


# controller.py (updated)
import pandas as pd
import numpy as np
from fastapi import UploadFile, HTTPException
from fastapi.responses import FileResponse
from app.services.groq_service import extract_sku_data
from app.utils.file_utils import read_pdf_text, save_excel
from openpyxl import load_workbook
import io
import os
import tempfile
from pathlib import Path

async def process_invoice(pdf_file: UploadFile, excel_file: UploadFile, vendor: str, temp_dir: str = None):
    """
    Process invoice - MUST return a dictionary, never a FileResponse
    Updated to use temp directory for better file management
    """
    try:
        print(f"🔄 Starting process_invoice for vendor: {vendor}")
        
        # Use provided temp directory or create one
        if temp_dir is None:
            temp_dir = tempfile.mkdtemp()
        
        # Step 1: Extract text from PDF
        print("📄 Extracting PDF text...")
        pdf_text = await read_pdf_text(pdf_file)
        print(f"✅ PDF text extracted: {len(pdf_text)} characters")
        
        # Step 2: Call Groq to get structured data
        print("🤖 Calling Groq service...")
        sku_data = extract_sku_data(pdf_text, vendor)
        
        if not sku_data:
            return {
                "status": "error",
                "message": "No SKU data extracted from PDF",
                "excel_data": [],
                "processed_skus": 0
            }
        
        print(f"✅ Extracted {len(sku_data)} SKUs")
        
        # Step 3: Update Excel quantities
        print("📊 Updating Excel quantities...")
        excel_data = update_excel_quantities(excel_file, sku_data, temp_dir)
        
        if not excel_data:
            return {
                "status": "error",
                "message": "Failed to update Excel file",
                "excel_data": [],
                "processed_skus": len(sku_data)
            }
        
        print(f"✅ Excel updated successfully. Records: {len(excel_data)}")
        
        # Return success result
        result = {
            "status": "success",
            "excel_data": excel_data,
            "message": f"Successfully processed {len(sku_data)} SKUs",
            "processed_skus": len(sku_data)
        }
        
        print(f"✅ Process completed successfully")
        return result
        
    except Exception as e:
        print(f"❌ Error in process_invoice: {str(e)}")
        print(f"📍 Error type: {type(e).__name__}")
        
        # Return error as dict, not raise exception
        return {
            "status": "error",
            "message": str(e),
            "excel_data": [],
            "processed_skus": 0
        }

def update_excel_quantities(excel_file: UploadFile, sku_data: dict, temp_dir: str) -> list:
    """
    Update Excel quantities and return data as list of dictionaries
    Updated to use temp directory for better file management
    """
    wb = None
    file_like = None
    
    try:
        print(f"📊 Starting Excel update with {len(sku_data)} SKUs")
        
        # Reset file pointer and read content
        excel_file.file.seek(0)
        file_content = excel_file.file.read()
        excel_file.file.seek(0)
        
        print(f"📁 File content size: {len(file_content)} bytes")
        
        # Create a BytesIO object for openpyxl
        file_like = io.BytesIO(file_content)
        
        # Load workbook from memory
        wb = load_workbook(file_like)
        print(f"📋 Workbook loaded. Sheets: {wb.sheetnames}")
        
        if "Items" not in wb.sheetnames:
            raise ValueError("'Items' sheet not found in Excel file")
        
        sheet = wb["Items"]
        print(f"✅ 'Items' sheet loaded")
        
        # Map header columns to index
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        header = {cell.value: idx for idx, cell in enumerate(header_row, start=1)}
        print(f"📋 Headers found: {list(header.keys())}")
        
        if 'SKU' not in header or 'Quantity' not in header:
            raise ValueError(f"Required columns not found. Available: {list(header.keys())}")
        
        sku_col = header['SKU']
        qty_col = header['Quantity']
        print(f"📍 SKU column: {sku_col}, Quantity column: {qty_col}")
        
        # Track updates
        updates_made = 0
        
        # Iterate over rows and update quantities
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            try:
                sku_cell = row[sku_col - 1]
                qty_cell = row[qty_col - 1]
                
                if sku_cell.value is None:
                    continue
                
                sku = str(sku_cell.value).strip()
                
                if sku in sku_data:
                    original_qty = qty_cell.value or 0
                    new_qty = original_qty + sku_data[sku]
                    qty_cell.value = new_qty
                    updates_made += 1
                    print(f"✅ Row {row_num} - SKU {sku}: {original_qty} + {sku_data[sku]} = {new_qty}")
                    
            except Exception as row_error:
                print(f"⚠️ Error processing row {row_num}: {row_error}")
                continue
        
        print(f"✅ Made {updates_made} updates to Excel file")
        
        # Save to temp directory
        output_path = os.path.join(temp_dir, "updated_invoice.xlsx")
        
        # Ensure temp directory exists
        Path(temp_dir).mkdir(parents=True, exist_ok=True)
        
        # Remove existing file if it exists
        if os.path.exists(output_path):
            os.remove(output_path)
            
        wb.save(output_path)
        print(f"💾 File saved to: {output_path}")
        
        # Verify file was created
        if not os.path.exists(output_path):
            raise FileNotFoundError(f"Failed to create output file: {output_path}")
            
        print(f"✅ File verification passed. Size: {os.path.getsize(output_path)} bytes")
        
        # Convert to DataFrame for returning data
        df = pd.read_excel(output_path, sheet_name="Items")
        df = df.replace({np.nan: None})  # Replace NaN with None for JSON serialization
        
        # Convert to list of dictionaries
        excel_data = df.to_dict(orient="records")
        print(f"📊 Returning {len(excel_data)} records")
        
        return excel_data
        
    except Exception as e:
        print(f"❌ Critical error in update_excel_quantities: {str(e)}")
        print(f"📍 Error type: {type(e).__name__}")
        return []
    
    finally:
        # Clean up file handles
        try:
            if wb:
                wb.close()
            if file_like:
                file_like.close()
        except Exception as cleanup_error:
            print(f"⚠️ Cleanup error (non-critical): {cleanup_error}")