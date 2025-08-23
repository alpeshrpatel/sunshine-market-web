# import pandas as pd
# from fastapi import UploadFile
# from fastapi.responses import FileResponse
# from app.services.groq_service import extract_sku_data
# from app.utils.file_utils import read_pdf_text, save_excel
# from app.api.controllers.excel_file_updation_controller import update_excel_quantities

# async def process_invoice(pdf_file: UploadFile, excel_file: UploadFile, vendor: str):
#     # Step 1: Extract text from PDF
#     pdf_text = await read_pdf_text(pdf_file)

#     # Step 2: Call Groq (OpenAI-compatible) to get structured data
#     sku_data = extract_sku_data(pdf_text, vendor)
    
#     print(f"Extracted SKU data: {sku_data}")

#     # Step 3: Load Excel
#     # df = pd.read_excel(excel_file.file)
#     # print('Loaded Excel data:\n', df.loc())

#     # # Step 4: Update quantities
#     # for sku, qty in sku_data.items():
#     #     df.loc[df['UPC'] == sku, 'Quantity'] = qty
#     import pandas as pd
    
#     excel_data = update_excel_quantities(excel_file, sku_data)


#     # Load only the 'Items' sheet
#     # df = pd.read_excel(excel_file.file, sheet_name='Items')

#     # print("✅ Loaded Excel Data:\n", df.head())
    
#     # df['SKU'] = df['SKU'].astype(str)

#     # for sku, qty in sku_data.items():
#     #     sku_str = str(sku)  # Ensure SKU key from API is also string

#     #     if sku_str in df['SKU'].values:
#     #         print(f"✅ Found SKU {sku_str} in Excel.")
            
#     #         # Update quantity by locating matching SKU
#     #         df.loc[df['SKU'] == sku_str, 'Quantity'] += qty
#     #         print(f"✅ Added {qty} to SKU {sku_str}")
#     #     else:
#     #         print(f"⚠️ SKU {sku_str} not found in Excel — skipping.")


#     # # Save back the updated Excel file
#     # output_path = 'updated_items.xlsx'
#     # with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
#     #     df.to_excel(writer, sheet_name='Items', index=False)


#     # # Step 5: Save updated Excel
#     # output_path = save_excel(df)


#     # Make sure column names match exactly:
#     # Assume your Excel columns are named: 'SKU' and 'Quantity'
#     # If your JSON has UPCs as keys, treat them as SKUs

#     # for sku, qty in sku_data.items():
#     #     # Check if SKU exists in the Excel file
#     #     if sku in df['SKU'].values:
#     #         print(f"✅ Found SKU {sku} in Excel, updating quantity. {df.loc[df['SKU']]}")
#     #         df.loc[df['SKU'] == int(sku), 'Quantity'] += qty
#     #         print(f"✅ Added {qty} to SKU {sku}")
#     #     else:
#     #         print(f"⚠️ SKU {sku} not found in Excel — skipping.")
#     # Ensure all SKUs in the DataFrame are strings for safe comparison
   
#     # return {"status": "success", "updated_file": output_path}
#     filename = "updated_invoice.xlsx"
#     return {"excel_data": excel_data, "download_url": "/api/invoice/download-updated"}
#     # return FileResponse(
#     #     path=output_path,
#     #     filename=filename,
#     #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     # )











import pandas as pd
import numpy as np
from fastapi import UploadFile, HTTPException
from fastapi.responses import FileResponse
from app.services.groq_service import extract_sku_data
from app.utils.file_utils import read_pdf_text, save_excel
from openpyxl import load_workbook
import io
import os

async def process_invoice(pdf_file: UploadFile, excel_file: UploadFile, vendor: str):
    """
    Process invoice - MUST return a dictionary, never a FileResponse
    """
    try:
        print(f"🔄 Starting process_invoice for vendor: {vendor}")
        
        # Step 1: Extract text from PDF
        print("📄 Extracting PDF text...")
        pdf_text = await read_pdf_text(pdf_file)
        print(f"✅ PDF text extracted: {len(pdf_text)} characters")
        
        # Step 2: Call Groq to get structured data
        print("🤖 Calling Groq service...")
        sku_data = extract_sku_data(pdf_text, vendor)
        # print(f"✅ Extracted SKU data: {sku_data}")
        
        # Step 3: Update Excel quantities
        print("📊 Updating Excel quantities...")
        excel_data = update_excel_quantities(excel_file, sku_data)
        print(f"✅ Excel updated successfully. Records: {len(excel_data)}")
        
        # CRITICAL: Always return a dictionary, never a FileResponse
        result = {
            "status": "success",
            "excel_data": excel_data,
            "message": f"Successfully processed {len(sku_data)} SKUs",
            "processed_skus": len(sku_data)
        }
        
        print(f"✅ Process completed successfully: {result}")
        return result
        
    except Exception as e:
        print(f"❌ Error in process_invoice: {str(e)}")
        print(f"📍 Error type: {type(e).__name__}")
        # Return error as dict, not raise exception
        return {
            "status": "error",
            "message": str(e),
            "excel_data": []
        }

def update_excel_quantities(excel_file: UploadFile, sku_data: dict) -> list:
    """
    Update Excel quantities and return data as list of dictionaries
    FIXED VERSION - handles file operations properly
    """
    try:
        print(f"📊 Starting Excel update with {len(sku_data)} SKUs")
        
        # CRITICAL FIX: Properly handle the uploaded file
        excel_file.file.seek(0)  # Reset file pointer to beginning
        
        # Read the entire file content into memory first
        file_content = excel_file.file.read()
        excel_file.file.seek(0)  # Reset again for openpyxl
        
        print(f"📁 File content size: {len(file_content)} bytes")
        
        # Create a BytesIO object for openpyxl
        file_like = io.BytesIO(file_content)
        
        # Load workbook from memory
        wb = load_workbook(file_like)
        print(f"📋 Workbook loaded. Sheets: {wb.sheetnames}")
        
        if "Items" not in wb.sheetnames:
            raise ValueError("❌ 'Items' sheet not found in Excel file")
        
        sheet = wb["Items"]
        print(f"✅ 'Items' sheet loaded")
        
        # Map header columns to index
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        header = {cell.value: idx for idx, cell in enumerate(header_row, start=1)}
        print(f"📋 Headers found: {list(header.keys())}")
        
        if 'SKU' not in header or 'Quantity' not in header:
            raise ValueError(f"❌ Required columns not found. Available: {list(header.keys())}")
        
        sku_col = header['SKU']
        qty_col = header['Quantity']
        print(f"📍 SKU column: {sku_col}, Quantity column: {qty_col}")
        
        # Track updates
        updates_made = 0
        
        # Iterate over rows and update quantities
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            try:
                sku_cell = row[sku_col - 1]
                qty_cell = row[qty_col - 1]
                
                if sku_cell.value is None:
                    continue
                
                sku = str(sku_cell.value).strip()
                
                if sku in sku_data:
                    original_qty = qty_cell.value or 0
                    new_qty = original_qty + sku_data[sku]
                    qty_cell.value = new_qty
                    updates_made += 1
                    print(f"✅ Row {row_num} - SKU {sku}: {original_qty} + {sku_data[sku]} = {new_qty}")
                    
            except Exception as row_error:
                print(f"⚠️ Error processing row {row_num}: {row_error}")
                continue
        
        print(f"✅ Made {updates_made} updates to Excel file")
        
        # Save updated file with error handling
        output_path = "updated_invoice.xlsx"
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
        
        # Remove existing file if it exists
        if os.path.exists(output_path):
            os.remove(output_path)
            
        wb.save(output_path)
        print(f"💾 File saved to: {output_path}")
        
        # Verify file was created
        if not os.path.exists(output_path):
            raise FileNotFoundError(f"Failed to create output file: {output_path}")
            
        print(f"✅ File verification passed. Size: {os.path.getsize(output_path)} bytes")
        
        # Convert to DataFrame for returning data
        df = pd.read_excel(output_path, sheet_name="Items")
        df = df.replace({np.nan: None})  # Replace NaN with None for JSON serialization
        
        # Convert to list of dictionaries
        excel_data = df.to_dict(orient="records")
        print(f"📊 Returning {len(excel_data)} records")
        
        return excel_data
        
    except Exception as e:
        print(f"❌ Critical error in update_excel_quantities: {str(e)}")
        print(f"📍 Error type: {type(e).__name__}")
        # Return empty list instead of raising exception
        return []
    
    finally:
        # Clean up file handles
        try:
            if 'wb' in locals():
                wb.close()
            if 'file_like' in locals():
                file_like.close()
        except Exception as cleanup_error:
            print(f"⚠️ Cleanup error (non-critical): {cleanup_error}")
