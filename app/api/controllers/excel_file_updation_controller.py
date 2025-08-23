from openpyxl import load_workbook
import pandas as pd
import numpy as np

def update_excel_quantities(excel_file, sku_data: dict) -> str:
    # Load workbook from uploaded file
    excel_file.file.seek(0)  # Reset file pointer
    wb = load_workbook(excel_file.file)
    
    if "Items" not in wb.sheetnames:
        raise Exception("❌ 'Items' sheet not found in Excel file")

    sheet = wb["Items"]

    # Map header columns to index
    header = {cell.value: idx for idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), start=1)}

    if 'SKU' not in header or 'Quantity' not in header:
        raise Exception("❌ 'SKU' or 'Quantity' column not found in 'Items' sheet")

    sku_col = header['SKU']
    qty_col = header['Quantity']

    # Iterate over rows and update quantities
    for row in sheet.iter_rows(min_row=2):
        sku_cell = row[sku_col - 1]
        qty_cell = row[qty_col - 1]

        sku = str(sku_cell.value).strip()
        if sku in sku_data:
            original_qty = qty_cell.value or 0
            qty_cell.value = original_qty + sku_data[sku]
            print(f"✅ SKU {sku}: {original_qty} + {sku_data[sku]} = {qty_cell.value}")

    # Save updated file
    output_path = "updated_invoice.xlsx"
    wb.save(output_path)
    
    df = pd.read_excel(output_path, sheet_name="Items")
    df = df.replace({np.nan: None})
    
    # Convert to dictionary
    # excel_data = df.to_dict(orient="records")
    excel_data = df.to_dict(orient="records")
    
    return excel_data
    
    # return output_path
